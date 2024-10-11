import sys
import os
from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QFileDialog, QLabel, QVBoxLayout, QWidget, QLineEdit, QHBoxLayout, QProgressDialog, QDialog, QTableWidget, QTableWidgetItem
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QDragEnterEvent, QDropEvent, QIcon
import subprocess
import pandas as pd
import warnings
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, NamedStyle

class Worker(QThread):
    progress = pyqtSignal(str)
    finished = pyqtSignal(bool, pd.DataFrame)
    error = pyqtSignal(str)

    def __init__(self, pdf_path, save_path, output_path):
        super().__init__()
        self.pdf_path = pdf_path
        self.save_path = save_path
        self.output_path = output_path

    def run(self):
        try:
            self.progress.emit('Конвертация PDF...')
            fs1_path = r'"C:\Program Files (x86)\ABBYY FineReader 15\FineCMD.exe"'
            work_dir = r'w:\_03_Вопросы_компании\Программы\FR\ '

            command = f'cmd /c cd /d "{work_dir}" && {fs1_path} "{self.pdf_path}" /out "{self.save_path}"'
            subprocess.run(command, shell=True, check=True)

            self.progress.emit('Обработка Excel файла...')
            dl_df = pd.read_excel(self.save_path)
            template_path = r"w:\_03_Вопросы_компании\УНФ\ЗагрузкаПоставщиков\исходники\Шаблон загрузки.xlsx"
            template_df = pd.read_excel(template_path, sheet_name="ДанныДляЗагрузки")
            headers = template_df.columns.tolist()
            ol_df = pd.DataFrame(columns=headers)

            last_row_dl = dl_df.shape[0]
            # Электропрофи
            start_row = self.get_start_row(dl_df, last_row_dl)
            if start_row > 0:
                ol_df = self.copy_data(dl_df, ol_df, start_row, last_row_dl)
            else:
                # Лицензии/Спецификация
                start_row = self.get_start_row_next(dl_df, last_row_dl)
                if start_row > 0:
                    ol_df = self.copy_data(dl_df, ol_df, start_row, last_row_dl, next=True)
                else:
                    # НТК Интерфейс
                    start_row = self.get_start_row_down(dl_df, last_row_dl)
                    if start_row > 0:
                        ol_df = self.copy_data(dl_df, ol_df, start_row, last_row_dl, down=True)
                    else:
                        # Электромонтаж
                        start_row = self.get_start_row_down_next(dl_df, last_row_dl)
                        if start_row > 0:
                            ol_df = self.copy_data(dl_df, ol_df, start_row, last_row_dl, down_next=True)
                        else:
                            self.error.emit("Неизвестный файл")
                            return

            self.finished.emit(True, ol_df)

        except subprocess.CalledProcessError as e:
            self.error.emit(f'Ошибка при конвертации PDF: {e}')
        except Exception as e:
            self.error.emit(f'Ошибка при обработке Excel файла: {e}')

    # Электропрофи
    def get_start_row(self, df, last_row_dl):
        for i in range(1, last_row_dl):
            value = df.iloc[i, 1]
            if isinstance(value, str) and value.lower() == "артикул":
                return i + 1
        return 0

    # Лицензии/Спецификация
    def get_start_row_next(self, df, last_row_dl):
        for i in range(1, last_row_dl):
            value = df.iloc[i, 1]
            if isinstance(value, str) and value.lower() == "товары (работы, услуги)":
                return i + 1
        return 0

    # Электротехмонтаж
    def get_start_row_down_next(self, df, last_row_dl):
        for i in range(1, last_row_dl):
            value = df.iloc[i, 1]
            if isinstance(value, str) and value.lower() == "код товара этм":
                return i + 1
        return 0

    # НТК Интерфейс
    def get_start_row_down(self, df, last_row_dl):
        for i in range(1, last_row_dl):
            value = df.iloc[i, 1]
            if isinstance(value, str) and value.lower() == "наименование":
                return i + 1
        return 0

    def copy_data(self, dl_df, ol_df, start_row, last_row_dl, next=False, down_next=False, down=False):
        new_rows = []
        for i in range(start_row, last_row_dl):
            # Прекращение копирования, если в столбце A есть слово "номенклатура"
            if isinstance(dl_df.iloc[i, 0], str) and dl_df.iloc[i, 0].lower() == "номенклатура":
                break
            # Электротехмонтаж
            if down_next:
                # Проверка, что значения в столбцах 1, 2, 3, 4 не являются только цифрами
                if all(not (isinstance(dl_df.iloc[i, col], (int, float)) or str(dl_df.iloc[i, col]).isdigit()) for col in range(1, 2)):
                    try:
                        price = float(str(dl_df.iloc[i, 7]).replace(' ', '').replace(',', '.'))
                        quantity = float(str(dl_df.iloc[i, 6]).replace('.', '').replace(' ', ''))
                    except ValueError:
                        continue  # Пропустить строки с некорректным значением цены

                    new_row = {
                        'Артикул поставщика': str(dl_df.iloc[i, 1]).replace(' ', ''),
                        'Номенклатура': str(dl_df.iloc[i, 2]).replace('двухп олюсный', 'двухполюсный'),
                        'Артикул': str(dl_df.iloc[i, 3]).replace(' ', ''),
                        'Количество': quantity,
                        'Единица измерения': dl_df.iloc[i, 5],
                        'Цена': price
                    }
                    new_rows.append(new_row)

            elif isinstance(dl_df.iloc[i, 0], str) and dl_df.iloc[i, 0].isdigit():
                if pd.isna(dl_df.iloc[i, 5]):
                    break
                try:
                    # Лицензии/Сертификаты
                    if next:
                        price = float(str(dl_df.iloc[i, 4]).replace(' ', '').replace(',', '.'))
                        quantity = float(str(dl_df.iloc[i, 2]).replace('.', '').replace(' ', ''))
                        new_row = {
                            'Номенклатура': dl_df.iloc[i, 1],
                            'Количество': quantity,
                            'Единица измерения': dl_df.iloc[i, 3],
                            'Цена': price
                        }
                    # НТК Интерфейс
                    elif down:
                        price = float(str(dl_df.iloc[i, 5]).replace(' ', '').replace(',', '.'))
                        quantity = float(str(dl_df.iloc[i, 4]).replace('.', '').replace(' ', ''))
                        new_row = {
                            'Номенклатура': dl_df.iloc[i, 2],
                            'Единица измерения': dl_df.iloc[i, 3],
                            'Количество': quantity,
                            'Цена': price
                        }
                    # Электропрофи
                    else:
                        price = float(str(dl_df.iloc[i, 5]).replace(' ', '').replace(',', '.'))
                        quantity = float(str(dl_df.iloc[i, 3]).replace('.', '').replace(' ', ''))
                        new_row = {
                            'Артикул': dl_df.iloc[i, 1],
                            'Номенклатура': dl_df.iloc[i, 2],
                            'Количество': quantity,
                            'Единица измерения': dl_df.iloc[i, 4],
                            'Цена': price
                        }
                    new_rows.append(new_row)
                except ValueError:
                    continue  # Пропустить строки с некорректным значением цены

        if new_rows:
            ol_df = pd.concat([ol_df, pd.DataFrame(new_rows, columns=ol_df.columns)], ignore_index=True)
        return ol_df

class PreviewDialog(QDialog):
    def __init__(self, preview_df, template_path, output_path, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Предпросмотр')
        self.setGeometry(350, 350, 1450, 600)
        self.output_path = output_path
        self.preview_df = preview_df
        self.setWindowIcon(QIcon('logo.ico'))
        
        # Считываем заголовки из шаблона
        self.template_headers = self.load_template_headers(template_path)
        
        layout = QVBoxLayout(self)
        
        self.table_widget = QTableWidget(self)
        self.table_widget.setRowCount(preview_df.shape[0])
        self.table_widget.setColumnCount(preview_df.shape[1])
        
        # Устанавливаем заголовки из шаблона
        self.table_widget.setHorizontalHeaderLabels(['Артикул поставщика', 'Артикул', 'Номенклатура', 'Характеристика', 'Категория', 'Количество', 'Единица измерения', 'Цена'])

        for i in range(preview_df.shape[0]):
            for j in range(preview_df.shape[1]):
                item = str(preview_df.iloc[i, j])
                item = '' if item == 'nan' else item
                if preview_df.columns[j] == 'Количество':
                    item = item.replace('.0', '').replace(',0', '')
                elif preview_df.columns[j] == 'Цена':
                    item = item.replace('.', ',')
                self.table_widget.setItem(i, j, QTableWidgetItem(item))

        # Автоматическая настройка ширины столбцов
        for j in range(preview_df.shape[1]):
            self.table_widget.resizeColumnToContents(j)


        layout.addWidget(self.table_widget)

        button_layout = QHBoxLayout()
        self.ok_button = QPushButton('Сохранить', self)
        self.close_button = QPushButton('Закрыть', self)  # Здесь заменили кнопку на "Закрыть"
        button_layout.addWidget(self.ok_button)
        button_layout.addWidget(self.close_button)  # Добавляем новую кнопку в layout

        # Добавление стилей к кнопкам
        button_style = """
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 10px;
                font-size: 14px;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """
        self.ok_button.setStyleSheet(button_style)
        self.close_button.setStyleSheet(button_style)
        
        self.ok_button.clicked.connect(self.save_changes)
        self.close_button.clicked.connect(self.accept)  # Закрывает диалог
        
        layout.addLayout(button_layout)
        self.setLayout(layout)


    def load_template_headers(self, template_path):
        try:
            template_df = pd.read_csv(template_path, header=0, encoding='utf-8')  # Указание кодировки
            return template_df.columns.tolist()
        except Exception as e:
            print(f"Ошибка при загрузке заголовков из шаблона: {e}")
            return []  # Возврат пустого списка в случае ошибки

        
    def open_document(self):
        if os.path.exists(self.output_path):
            os.startfile(self.output_path)
    
    def save_changes(self):
        for i in range(self.table_widget.rowCount()):
            for j in range(self.table_widget.columnCount()):
                item = self.table_widget.item(i, j).text()
                if self.preview_df.columns[j] == 'Количество':
                    item = item.replace(',', '.').replace(' ', '')  # Удалить пробелы и запятые
                    try:
                        item = float(item)
                    except ValueError:
                        item = ''
                elif self.preview_df.columns[j] == 'Цена':
                    item = item.replace(' ', '').replace(',', '.')  # Заменить запятые на точки
                    try:
                        item = float(item)
                    except ValueError:
                        item = ''
                self.preview_df.iloc[i, j] = item
        self.accept()

class PDFtoExcelConverter(QMainWindow):
    def __init__(self):
        super().__init__()

        self.initUI()
        self.apply_styles()

    def initUI(self):
        self.setWindowTitle('Конвертер PDF в Excel')
        self.setGeometry(800, 300, 400, 350)
        self.setAcceptDrops(True)
        self.setFixedWidth(600)
        self.setWindowIcon(QIcon('logo.ico'))

        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)

        layout = QVBoxLayout(self.central_widget)

        self.label = QLabel('Перетащите PDF файл сюда или используйте кнопку ниже', self)
        self.label.setWordWrap(True)
        layout.addWidget(self.label)

        self.save_path_label = QLabel('Путь сохранения конвертируемого PDF:', self)
        layout.addWidget(self.save_path_label)

        self.save_path_input = QLineEdit(self)
        self.save_path_input.setPlaceholderText(r'w:\_03_Вопросы_компании\УНФ\ЗагрузкаПоставщиков\исходники\converted.xlsx')
        layout.addWidget(self.save_path_input)

        self.output_path_label = QLabel('Путь сохранения Excel таблицы:', self)
        layout.addWidget(self.output_path_label)

        self.output_path_input = QLineEdit(self)
        self.output_path_input.setPlaceholderText(r'w:\_03_Вопросы_компании\УНФ\ЗагрузкаПоставщиков\Загрузка1.xlsx')
        layout.addWidget(self.output_path_input)

        hbox = QHBoxLayout()
        layout.addLayout(hbox)

        self.btn_select_pdf = QPushButton('Выбрать PDF', self)
        self.btn_select_pdf.clicked.connect(self.select_pdf)
        hbox.addWidget(self.btn_select_pdf)

        self.btn_convert = QPushButton('Конвертировать в Excel', self)
        self.btn_convert.clicked.connect(self.start_conversion)
        hbox.addWidget(self.btn_convert)

        self.btn_clear = QPushButton('Очистить поля', self)
        self.btn_clear.clicked.connect(self.clear_fields)
        hbox.addWidget(self.btn_clear)

        self.pdf_path = None
        self.default_save_path = r"w:\_03_Вопросы_компании\УНФ\ЗагрузкаПоставщиков\исходники\converted.xlsx"
        self.default_output_path = r"w:\_03_Вопросы_компании\УНФ\ЗагрузкаПоставщиков\Загрузка1.xlsx"

    def apply_styles(self):
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f0f0f0;
            }
            QLabel {
                color: #333;
                font-size: 18px;
            }
            QLineEdit {
                border: 1px solid #ccc;
                border-radius: 4px;
                padding: 8px;
                font-size: 14px;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 10px;
                font-size: 14px;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:pressed {
                background-color: #388E3C;
            }
            QVBoxLayout, QHBoxLayout {
                padding: 10px;
            }
            QWidget {
                padding: 10px;
            }
            QTableWidget {
                background-color: #ffffff;
                border: 1px solid #ccc;
                gridline-color: #dddddd;
            }
            QTableWidget::item {
                padding: 8px;
                border: 1px solid #e0e0e0;
            }
            QTableWidget::item:selected {
                background-color: #4CAF50;
                color: white;
            }
            QHeaderView::section {
                background-color: #4CAF50;
                color: white;
                padding: 5px;
                font-size: 16px; /* Увеличенный размер шрифта */
                font-weight: bold; /* Жирный шрифт */
            }
        """)


    def select_pdf(self):
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly
        file_name, _ = QFileDialog.getOpenFileName(self, "Выберите PDF файл", "", "PDF Files (*.pdf)", options=options)
        if file_name:
            self.pdf_path = file_name
            self.label.setText(f'Выбранный PDF: {self.pdf_path}')

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event: QDropEvent):
        urls = event.mimeData().urls()
        if urls and urls[0].isLocalFile():
            self.pdf_path = urls[0].toLocalFile()
            self.label.setText(f'Выбранный PDF: {self.pdf_path}')

    def start_conversion(self):
        if not self.pdf_path:
            self.label.setText('Пожалуйста, выберите PDF файл')
            return

        save_path = self.save_path_input.text()
        if not save_path:
            save_path = self.default_save_path
            self.save_path_input.setText(save_path)

        output_path = self.output_path_input.text()
        if not output_path:
            output_path = self.default_output_path
            self.output_path_input.setText(output_path)

        self.progress_dialog = QProgressDialog("Выполнение операции...", "Отмена", 0, 0, self)
        self.progress_dialog.setWindowModality(Qt.WindowModal)

        # Подключаем кнопку "Отмена" к методу
        self.progress_dialog.canceled.connect(self.cancel_conversion)  
        self.progress_dialog.show()

        self.worker = Worker(self.pdf_path, save_path, output_path)
        self.worker.progress.connect(self.update_progress)
        self.worker.finished.connect(self.conversion_finished)
        self.worker.error.connect(self.show_error)
        self.worker.start()

    def update_progress(self, message):
        self.progress_dialog.setLabelText(message)

    def conversion_finished(self, success, df):
        self.progress_dialog.close()
        if success:
            self.label.setText('Данные успешно скопированы и сохранены!')
            self.show_preview(df)
        else:
            self.label.setText('Ошибка при выполнении операции')

    def show_error(self, message):
        self.progress_dialog.close()
        self.label.setText(message)

    def clear_fields(self):
        self.pdf_path = None
        self.label.setText('Перетащите PDF файл сюда или используйте кнопку ниже')
        self.save_path_input.clear()
        self.output_path_input.clear()

    def show_preview(self, df):
        output_path = self.output_path_input.text()
        preview_dialog = PreviewDialog(df, output_path, self)
        if preview_dialog.exec_() == QDialog.Accepted:
            self.save_excel_file(df, output_path)

    def save_excel_file(self, df, output_path):
        wb = Workbook()
        ws = wb.active
        default_style = NamedStyle(name="default")
        default_style.font = Font(name='Arial', size=12)
        wb.add_named_style(default_style)

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            ws.append(row)
            if r_idx == 1:
                for cell in ws[r_idx]:
                    cell.style = default_style
                    cell.font = Font(bold=True)
            else:
                for cell in ws[r_idx]:
                    cell.style = default_style

        wb.save(output_path)

    def cancel_conversion(self):
        # Здесь вы можете выполнить любые необходимые действия при отмене
        self.label.setText('Операция отменена пользователем. ' 'Добавьте PDF файл')
        # Если у вас есть доступ к вашему Worker, можете также остановить его
        if self.worker:
            self.worker.terminate()  # Прекращаем работу воркера, если это возможно

if __name__ == '__main__':
    warnings.filterwarnings("ignore")
    app = QApplication(sys.argv)
    ex = PDFtoExcelConverter()
    ex.show()
    sys.exit(app.exec_())
